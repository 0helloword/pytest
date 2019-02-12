#coding:utf-8
#import random

from openpyxl import Workbook, load_workbook
from contextlib import closing


import sys
reload(sys)


class recordLoanData(object):
    ROBOT_LIBRARY_SCOPE = 'GLOBAL'
    ROBOT_LIBRARY_VERSION = 0.1


    def __init__(self):
        pass

    ## 创建Excel
    def createExcel(self, file_name):
        with closing(Workbook()) as wb:
            wb.save(file_name)
            
    
    ## 获取Excel行数
    def getExcelRowCount(self, file_name, sheet_name):
        with closing(load_workbook(filename=file_name)) as wb:
            rows = wb.get_sheet_by_name(name = str(sheet_name)).max_row
            return rows
    ## 获取Excel列数
    def getExcelColumnCount(self, file_name, sheet_name):
        with closing(load_workbook(filename=file_name)) as wb:
            columns = wb.get_sheet_by_name(name = str(sheet_name)).max_column
            return columns
            
    ## 读取从某一行某一列开始的n个值
    def listReadExcel(self, file_name, sheet_name, cell_row,cell_col,n ):
        with closing(load_workbook(filename=file_name, data_only=True)) as wb:
            ws = wb.get_sheet_by_name(str(sheet_name))
            cellValueList = []
            for i in range(int(n)):
                cellValue = ws.cell(row=int(cell_row),column=int(cell_col)+i).value
                cellValueList.append(cellValue)
            return cellValueList
            
    ## 读取从某一列某一行开始的n个值
    def listExcel(self, file_name, sheet_name, cell_row, cell_col, n):
        with closing(load_workbook(filename=file_name, data_only=True)) as wb:
            ws = wb.get_sheet_by_name(str(sheet_name))
            cellValueList = []
            for i in range(int(n)):
                cellValue = ws.cell(row=int(cell_row)+i,column=int(cell_col)).value
                cellValueList.append(cellValue)
            return cellValueList